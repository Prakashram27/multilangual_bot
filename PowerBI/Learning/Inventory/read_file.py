import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Prakash\\Downloads\\Laptops.xlsx")

dataframe1 = book.active

for row in range(0,dataframe1.max_row):
    for col in dataframe1.iter_cols(1,dataframe1.max_column):
        print(col[row].value)